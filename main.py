from openpyxl import *
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
import time

# selenium driver and webportal login details
USERNAME = "DCB10103"
PASSWORD = "*******"
ser = Service("/Users/satyaprakashsahoo/Documents/Chrome Driver/chromedriver")
driver = webdriver.Chrome(service=ser)
driver.get("https://dcb-spielerpass.de/")


# In the excel separate out the first name and last name into 2 different columns
# Only keep the new names to be registered
# Store images as firstname.jpg in image folder
# Store declaration as firstname.jpg in declaration folder
# read from excel and return player details
def get_player(player):
    workbook = load_workbook('/Users/satyaprakashsahoo/Library/Mobile '
                             'Documents/com~apple~CloudDocs/Personal/Cricket/BC '
                             'Attaching/Spielerpass/2022/Playerpass/PP_up.xlsx')
    sheet = workbook.active
    email = sheet.cell(row=player, column=2).value
    first_name = sheet.cell(row=player, column=3).value.replace(" ", "")
    last_name = sheet.cell(row=player, column=4).value
    date_birth = sheet.cell(row=player, column=5).value
    date_birth = date_birth.strftime("%d.%m.%Y")
    since_germany = sheet.cell(row=player, column=6).value
    since_germany = since_germany.strftime("%d.%m.%Y")
    nationality = sheet.cell(row=player, column=7).value
    id_number = sheet.cell(row=player, column=8).value
    address = sheet.cell(row=player, column=9).value
    address_split = address.split(" ")
    street = address_split[0].replace(",", "").replace(".", "")
    apartment = address_split[1].replace(",", "").replace(".", "")
    postcode = address_split[2].replace(",", "").replace(".", "")
    city = address_split[3].replace(",", "").replace(".", "")
    return [email, first_name, last_name, date_birth, since_germany, nationality, id_number, street, apartment,
            postcode, city]


# login to dcb portal
username = driver.find_element(By.ID, "user_login")
username.send_keys(USERNAME)
password = driver.find_element(By.ID, "user_pass")
password.send_keys(PASSWORD)
password.send_keys(Keys.ENTER)
time.sleep(2)

# register new players. Decide the range based on players to be entered. Start with 2.
# image cropping is a manual step, cannot be automated. Crop within 10 seconds to prevent error.
for n in range(2, 4):
    player_det = get_player(n)
    register_button = driver.find_element(By.XPATH,
                                          '//*[@id="post-25"]/div/div/section[4]/div/div[1]/div/div/div/div/a[2]')
    register_button.click()
    time.sleep(2)
    first_name_field = driver.find_element(By.ID, "firstname")
    first_name_field.send_keys(player_det[1])
    last_name_field = driver.find_element(By.ID, "lastname")
    last_name_field.send_keys(player_det[2])
    birthdate_field = driver.find_element(By.ID, "birthdate")
    birthdate_field.send_keys(player_det[3])
    time.sleep(2)
    ingermany_field = driver.find_element(By.ID, "ingermany")
    ingermany_field.send_keys(player_det[4])
    time.sleep(2)
    nationality_field = driver.find_element(By.ID, "nationality")
    nationality_field.send_keys(player_det[5])
    email_field = driver.find_element(By.ID, "playeremail")
    email_field.send_keys(player_det[0])
    id_number_field = driver.find_element(By.ID, "idproofnumber")
    id_number_field.send_keys(player_det[6])
    authority_field = driver.find_element(By.ID, "approvingauthority")
    authority_field.send_keys("Indian Consulate, Munich")
    photo_button = driver.find_element(By.ID, "photo")
    photo_button.send_keys(f"/Users/satyaprakashsahoo/Library/Mobile "
                           f"Documents/com~apple~CloudDocs/Personal/Cricket/BC "
                           f"Attaching/Spielerpass/2022/Playerpass/Pictures/{player_det[1]}.jpg")
    time.sleep(20)
    declaration_button = driver.find_element(By.ID, "idproof")
    declaration_button.send_keys(f"/Users/satyaprakashsahoo/Library/Mobile "
                                 f"Documents/com~apple~CloudDocs/Personal/Cricket/BC "
                                 f"Attaching/Spielerpass/2022/Playerpass/Declaration/{player_det[1]}.jpg")
    time.sleep(30)
    next_1_button = driver.find_element(By.XPATH, '//*[@id="btn_personal_details"]')
    next_1_button.click()
    time.sleep(2)
    street_field = driver.find_element(By.ID, "street")
    street_field.send_keys(player_det[7])
    apartment_field = driver.find_element(By.ID, "apartmentnumber")
    apartment_field.send_keys(player_det[8])
    city_field = driver.find_element(By.ID, "city")
    city_field.send_keys(player_det[10])
    zip_field = driver.find_element(By.ID, "zip")
    zip_field.send_keys(player_det[9])
    time.sleep(20)
    next_2_button = driver.find_element(By.XPATH, '// *[ @ id = "btn_address_details"]')
    next_2_button.click()
    time.sleep(2)
    next_3_button = driver.find_element(By.XPATH, '// *[ @ id = "btn_club_details"]')
    next_3_button.click()
    time.sleep(2)
    submit_field = driver.find_element(By.ID, "submitnewplayer")
    submit_field.click()
    admin_page = driver.find_element(By.XPATH, '//*[@id="post-48"]/div/div/section[3]/div/div/div/div/div/div/a')
    admin_page.click()
    time.sleep(5)
